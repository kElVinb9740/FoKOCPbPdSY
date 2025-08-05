# 代码生成时间: 2025-08-05 11:57:23
import asyncio
from sanic import Sanic, response
from sanic.request import Request
from sanic.exceptions import ServerError
from openpyxl import Workbook
import os
from openpyxl.utils.exceptions import InvalidFileError

# 定义一个简单的Sanic应用
app = Sanic("ExcelGenerator")

# 路由 /generate_excel, 用于生成Excel文件
@app.route("/generate_excel", methods=["GET"])
async def generate_excel(request: Request):
    # 获取查询参数
    try:
        data = request.args.get('data')
        if not data:
            return response.json({
                "error": "Missing data parameter"
            }, status=400)
        
        # 解析data参数，这里假设data是以逗号分隔的多个列数据
        columns = data.split(',')
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        
        # 将列名写入Excel的第一行
        for col, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col).value = col_name
        
        # 假设生成的数据只有一行，这里可以根据需要修改
        # 将数据写入Excel的第二行
        for col, value in enumerate(columns, start=1):
            ws.cell(row=2, column=col).value = value
        
        # 保存Excel文件
        file_name = "generated_excel.xlsx"
        wb.save(filename=file_name)
        
        # 将文件作为响应发送给客户端
        return response.file(file_name)
    except InvalidFileError:
        return response.json({
            "error": "Invalid Excel file"
        }, status=500)
    except Exception as e:
        raise ServerError(f"An error occurred: {e}", status=500)

# 运行Sanic应用
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000, auto_reload=False)
